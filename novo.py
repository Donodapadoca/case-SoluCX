
import openpyxl
import pandas as pd


dados = pd.read_excel('clientes1.xlsx', sheet_name='clientes')
dados1 = pd.read_excel('auto.xlsx',sheet_name='auto')

workbook = openpyxl.Workbook()
planilha = workbook.active
dadospestado = dados.groupby('state')

#cria uma nova aba no ecxel
nsheet = workbook.create_sheet("Cidades - RS")

#pega a coluna monthly_income e "salva" em salario
salario = 'monthly_income'
#pega a coluna state e "salva" em estados
estados = dados['state']



#separa os estados do sudeste
estadossud = ['SP', 'ES', 'MG','RJ']

numero_coluna = 2


estsud_data = estados[estados.isin(estadossud)]
#agrupa e faz a media dos salarios dos estados do sudeste
salariopestado = dados.groupby(estsud_data)[salario].mean()


# Colocando a string Estado e Média em frente aos estados do sudeste
est = planilha.cell(row=1, column=4)
med = planilha.cell(row=1, column=5)
est.value = "Estados"
med.value = "Média"


#organiza as medias e os estados em linha por linha e coluna por coluna( apenas os estados do Sudeste)

for estado, mediasalario in salariopestado.items():
    planilha.cell(row = numero_coluna, column = 4).value = estado
    planilha.cell(row = numero_coluna, column = 5).value = mediasalario
    numero_coluna += 2



#pega os dados dos estados, e compara o menor e o maior para imprimir
maiorpestado = dadospestado['monthly_income'].max()
menorpestado = dadospestado['monthly_income'].min()



# planilha.cell(row = 1, column = 1).value = "Estado"
# planilha.cell(row = 1, column = 2).value = "Salario"






#coloca strings em frente aos dados para melhor entendimento
est = planilha.cell(row=1,column=7)
maior = planilha.cell(row=1, column=8)
menor = planilha.cell(row=1, column=11)
est.value = "Estados"
maior.value = "Maior"
menor.value = "Menor"

numero_coluna = 2
#organiza os maiores salarios por estado
for estado, maioestado in maiorpestado.items():
    planilha.cell(row = numero_coluna, column=7).value = estado
    planilha.cell(row = numero_coluna, column=8).value = maioestado
    numero_coluna +=1


est = planilha.cell(row=1,column=10)
est.value = "Estados"
numero_coluna = 2
#organiza os menores salarios por estado
for estado, menorestado in menorpestado.items():
    planilha.cell(row= numero_coluna, column=10).value = estado
    planilha.cell(row=numero_coluna, column=11).value = menorestado
    numero_coluna += 1


p60mais = (dados['age']>60).mean() * 100

pformat = f"{p60mais:.2f}%"

mais60 = planilha.cell(row=1, column=14)
mais60.value = "Clientes 60+"
planilha ['N2'].value = pformat

#combina os" dados da planilha clientes com a planilha auto 
#utiliza o "how='left'" para garantir que todas as informaçoes dos clientes sejam mantidas, mesmo aqueles que não possuem carro
dadoscombinados = dados.merge(dados1, how='left', on='id_cliente')





#encontrado os carros com a marca ford
dadosford = dadoscombinados.query("auto_brand== 'Ford' ")

#agrupando os carros com a marca ford em cada cidade x utilizando o id_cliente
cidadecarro = dadosford.groupby('city')['id_cliente'].size()

cmaiscarros = cidadecarro.idxmax()
mcontagem = cidadecarro.max()

cidade = planilha.cell(row=1,column=16)
qtd = planilha.cell(row=1, column=17)
cidade.value = "Cidade"
qtd.value = "Quantidade"

#imprimindo na planiha a cidade que contém mais carros da marca Ford
city = planilha.cell(row=2, column=16)
qts = planilha.cell(row=2, column=17)
city.value = cmaiscarros
qts.value = mcontagem



#procurar clientes pertencentes a cidades do estado do RS
# estadosrs = dados.query("state == 'RS' ")

#limpar todas as cidades duplicadas que provavelmente vão existir e contar elas

# cunicas = estadosrs.drop_duplicates(subset='city')
# ncidades = cunicas.shape[0]



# ccoluna = 4
# scoluna = 5
# filtro_cidades = "RS"

# stategup = {}

# for row in planilha.iter_rows(min_row = 2):
#     cnome = row[ccoluna - 1].value
#     snome = row[scoluna - 1].value


#     if snome == filtro_cidades:
#         if snome not in stategup:
#             stategup[snome] = []
#         stategup[snome].append(cnome)



# for snome, cities in stategup.items():
#     print(f"\nCidades do estado {snome}:")

#     for ctt in sorted(set(cities)):
#         print(ctt)





# print(rsestado)




















workbook.save('Dadosfinal.xlsx')





