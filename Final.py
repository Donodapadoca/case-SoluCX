
import openpyxl
import pandas as pd


dados = pd.read_excel('clientes1.xlsx', sheet_name='clientes')
dados1 = pd.read_excel('auto.xlsx',sheet_name='auto')

workbook = openpyxl.Workbook()
planilha = workbook.active
dadospestado = dados.groupby('state')
Nova_plan = workbook.create_sheet("Cidades - RS")

#pega a coluna monthly_income e "salva" em salario
salario = 'monthly_income'
#pega a coluna state e "salva" em estados
estados = dados['state']

# removendo os espacos em branco
dados["city"] = dados["city"].str.strip()

#convertendo todas as cidades para maiusculo para ver se o filtro de drop_duplicates funciona melhor
dados["city"] = dados["city"].str.upper()

# Filtragem por estado
Cidades_rs = dados[dados["state"] == "RS"]

#separa os estados do sudeste
estadossud = ['SP', 'ES', 'MG','RJ']

Numero_linha = 2


estsud_data = estados[estados.isin(estadossud)]
#agrupa e faz a media dos salarios dos estados do sudeste
salariopestado = dados.groupby(estsud_data)[salario].mean()


# Colocando a string Estado e Média em frente aos estados do sudeste
est = planilha.cell(row=1, column=4)
med = planilha.cell(row=1, column=5)
est.value = "Estados"
med.value = "Média"




#organiza as medias e os estados em linha por linha e coluna por coluna( apenas os estados do Sudeste)
for estado, mediasalario in salariopestado.items():
    planilha.cell(row = Numero_linha, column = 4).value = estado
    planilha.cell(row = Numero_linha, column = 5).value = mediasalario
    Numero_linha += 2



#pega os dados dos estados, e compara o menor e o maior para imprimir
maiorpestado = dadospestado['monthly_income'].max()
menorpestado = dadospestado['monthly_income'].min()



#coloca strings em frente aos dados para melhor entendimento
est = planilha.cell(row=1,column=7)
maior = planilha.cell(row=1, column=8)
menor = planilha.cell(row=1, column=11)
est.value = "Estados"
maior.value = "Maior"
menor.value = "Menor"

Numero_linha = 2


#organiza os maiores salarios por estado
for estado, maioestado in maiorpestado.items():
    planilha.cell(row = Numero_linha, column=7).value = estado
    planilha.cell(row = Numero_linha, column=8).value = maioestado
    Numero_linha +=1


est = planilha.cell(row=1,column=10)
est.value = "Estados"
Numero_linha = 2

#organiza os menores salarios por estado
for estado, menorestado in menorpestado.items():
    planilha.cell(row= Numero_linha, column=10).value = estado
    planilha.cell(row=Numero_linha, column=11).value = menorestado
    Numero_linha += 1


p60mais = (dados['age']>60).mean() * 100

pformat = f"{p60mais:.2f}%"

mais60 = planilha.cell(row=1, column=14)
mais60.value = "Clientes 60+"
planilha ['N2'].value = pformat

#combina os" dados da planilha clientes com a planilha auto 
#utiliza o "how='left'" para garantir que todas as informaçoes dos clientes sejam mantidas, mesmo aqueles que não possuem carro
dadoscombinados = dados.merge(dados1, how='left', on='id_cliente')


#encontrado os carros com a marca ford
dadosford = dadoscombinados.query("auto_brand== 'Ford' ")

#agrupando os carros com a marca ford em cada cidade x utilizando o id_cliente
cidadecarro = dadosford.groupby('city')['id_cliente'].size()

cmaiscarros = cidadecarro.idxmax()
mcontagem = cidadecarro.max()

cidade = planilha.cell(row=1,column=16)
qtd = planilha.cell(row=1, column=17)
cidade.value = "Cidade"
qtd.value = "Quantidade"

#imprimindo na planiha a cidade que contém mais carros da marca Ford
city = planilha.cell(row=2, column=16)
qts = planilha.cell(row=2, column=17)
city.value = cmaiscarros
qts.value = mcontagem


# Filtragem por estado
Cidades_rs = dados[dados["state"] == "RS"]

# Remoção das duplicatas
Cidadesunicas_rs = Cidades_rs.drop_duplicates(subset="city")

#ignorando as linhas em branco
Sembranco_cidades = Cidadesunicas_rs.dropna(subset=["city"])

# Contagem de cidades únicas
Qtd_cidades = len(Sembranco_cidades)

Qtd_rs = Nova_plan.cell(row = 1, column = 1)
Qtd_rs.value = "Qtd de Cidades"


N_cidades_rs = Nova_plan.cell(row = 1, column = 4)
N_cidades_rs.value = "Cidades RS"

# Escrita na planilha "Cidades - RS"
QT_cidades = Nova_plan.cell(row=2, column=1)
QT_cidades.value = Qtd_cidades

#imprime as cidades na nova aba criada
for i, cidade in enumerate(Sembranco_cidades["city"], start=2):
    Nova_plan.cell(row=i, column=4).value = cidade


print(Qtd_cidades)


#salva os dados na planilha escolhida
workbook.save('Dadosfinal.xlsx')




