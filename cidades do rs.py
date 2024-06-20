# import pandas as pd
# import openpyxl
# from openpyxl import Workbook




# dados = pd.read_excel("clientes1.xlsx",sheet_name='clientes')
# wb = openpyxl.Workbook()
# ws = wb.active

# Nova_plan = wb.create_sheet("Cidades - RS")



# cidades_rs = dados[dados["state"] == "RS"]

# cidades_rs["cidade_maiuscula"] = cidades_rs["city"].apply(lambda x: x.upper())



# cidades_rs_unicas = cidades_rs.drop_duplicates(subset=["cidade_maiuscula"])

# numero_cidades_rs = len(cidades_rs_unicas)


# print(f"Número de cidades do RS:", numero_cidades_rs)



# Qtd_cidades = ws.cell(row = 2, column = 1)
# Qtd_cidades.value = numero_cidades_rs


# for i, cidade in enumerate(cidades_rs_unicas["cidade_maiuscula"], start=2):
#     Nova_plan.cell(row=i, column=2).value = cidade


# print(numero_cidades_rs)


# wb.save("testedados.xlsx")


import pandas as pd
import openpyxl
from openpyxl import Workbook

# Leitura da tabela
dados = pd.read_excel("clientes1.xlsx", sheet_name="clientes")

# Criação do workbook e das planilhas
wb = openpyxl.Workbook()
ws = wb.active
Nova_plan = wb.create_sheet("Cidades - RS")

# Filtragem por estado
cidades_rs = dados[dados["state"] == "RS"]

# Conversão para maiúsculas e remoção de duplicatas
cidades_rs["cidade_maiuscula"] = cidades_rs["city"].apply(lambda x: x.upper())
cidades_rs_unicas = cidades_rs.drop_duplicates(subset=["cidade_maiuscula"])

# Contagem de cidades únicas
numero_cidades_rs = len(cidades_rs_unicas)

# Escrita na planilha "Cidades - RS"
Qtd_cidades = ws.cell(row=2, column=1)
Qtd_cidades.value = numero_cidades_rs

for i, cidade in enumerate(cidades_rs_unicas["cidade_maiuscula"], start=2):
    Nova_plan.cell(row=i, column=2).value = cidade

# Salvamento do arquivo Excel
wb.save("testedados.xlsx")

print(f"Número de cidades do RS:", numero_cidades_rs)
